import json

import django_filters
from django.db.models import Q
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from rest_framework import generics, viewsets
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import *
from .serializers import *


class ObjectFilter(django_filters.FilterSet):
    q = django_filters.CharFilter(method='my_custom_filter', label="Search")

    def my_custom_filter(self, queryset, name, value):
        return queryset.filter(
            Q(number__icontains=value) |
            Q(id__icontains=value) |
            Q(name__icontains=value) |
            Q(address__icontains=value)
        )

    class Meta:
        model = Object
        fields = ['name']

class ObjectPagination(PageNumberPagination):
    page_size = 15
    page_size_query_param = 'page_size'
    max_page_size = 10000

class ObjectViewSet(viewsets.ModelViewSet):
    pagination_class = ObjectPagination
    queryset = Object.objects.all()
    serializer_class = ObjectSerializer
    lookup_field = 'id'
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = ObjectFilter

    def create(self, request, *args, **kwargs):
        print(request.data)
        setattr(request.data, '_mutable', True)

        try:
            request.data.pop('image')
        except:
            pass
        try:
            request.data.pop('files')
            files_descriptions = request.data.pop('descriptions')
        except:
            files_descriptions = []
        data = json.loads(json.dumps(request.data))
        json_data = {}
        for dat in data:
            json_data[dat] = json.loads(data[dat])
        serializer = self.get_serializer(data=json_data)
        if serializer.is_valid():
            obj = serializer.save()
            obj.client_id = json_data['client']

            image = request.FILES.get('image',None)
            if image:
                obj.image = image
            obj.save()
            for index,file in enumerate(request.FILES.getlist('files')):
                ObjectFile.objects.create(file=file,object=obj,text=files_descriptions[index])

            for contact in json_data['contacts']:
                contact_serializer = ObjectContactSerializer(data=contact)
                if contact_serializer.is_valid():
                    contact_obj = contact_serializer.save()
                    contact_obj.object = obj
                    contact_obj.save()
            for equipment in json_data['equipments']:
                ObjectAdditionalEquipment.objects.create(object=obj,
                                                         model_id=equipment['model'],
                                                         amount=equipment['amount'])
        else:
            print(serializer.errors)

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class GetAddEqCategory(generics.ListAPIView):
    serializer_class = AdditionalEquipmentCategorySerializer
    queryset = AdditionalEquipmentCategory.objects.all()

class GetAddEqModel(generics.ListAPIView):
    serializer_class = AdditionalEquipmentModelSerializer
    def get_queryset(self):
        return AdditionalEquipmentModel.objects.filter(category_id=self.request.query_params.get('c_id'))




class DeleteAddEq(APIView):
    def post(self,request):

        obj = ObjectAdditionalEquipment.objects.get(id=request.data['e_id'])
        obj.delete()
        return Response(status=200)

class DeleteFile(APIView):
    def post(self,request):
        print(request.data)
        obj = ObjectFile.objects.get(id=request.data['f_id'])
        obj.delete()
        return Response(status=200)

class DeleteContact(APIView):
    def post(self,request):
        print(request.data)
        obj = ObjectContact.objects.get(id=request.data['c_id'])
        obj.delete()
        return Response(status=200)

class AddEqip(APIView):
    def get(self, request):
        from openpyxl import load_workbook
        from equipment.models import Equipment,EquipmentModel
        wb = load_workbook(filename='equip.xlsx')
        sheet_obj = wb.active
        max_row = sheet_obj.max_row
        for i in range(2, max_row + 3):
            obj_number = sheet_obj.cell(row=i, column=1)
            if obj_number.value:
                obj = Object.objects.filter(number=obj_number.value).first()
                if obj:
                    model_id = sheet_obj.cell(row=i, column=2)
                    equip_serial = sheet_obj.cell(row=i, column=3)
                    equip_pay_date = sheet_obj.cell(row=i, column=4)
                    equip_date_in_work = sheet_obj.cell(row=i, column=5)
                    equip_service_book_date = sheet_obj.cell(row=i, column=6)
                    print(obj_number.value, model_id.value,equip_serial.value,equip_pay_date.value, equip_date_in_work.value, equip_service_book_date.value)
                    if equip_serial.value :
                        try:
                            equipment, created = Equipment.objects.get_or_create(serial_number=equip_serial.value)
                            if created:
                                equipment.model_id = model_id.value
                            if equip_date_in_work.value and equip_date_in_work.value != ' ':
                                equipment.date_in_work = equip_date_in_work.value
                            if equip_pay_date.value and equip_pay_date.value != ' ':
                                equipment.pay_date = equip_pay_date.value
                            if equip_service_book_date.value and equip_service_book_date.value != ' ':
                                equipment.service_book_sign_date = equip_service_book_date.value
                                equipment.is_service_book_sign = True

                            equipment.object = obj
                            equipment.save()
                        except:
                            print('error')
                else:
                    print('no object', obj_number.value)
    #             print(obj_number.value,
    # model_id.value,
    # equip_serial.value,
    # equip_date_in_work.value,
    # equip_service_book_date.value)


        return Response(status=200)
class FillObject(APIView):

    def get(self,request):
        from openpyxl import load_workbook
        wb = load_workbook(filename='obj2.xlsx')
        sheet_obj = wb.active
        max_row = sheet_obj.max_row

        for i in range(5,  max_row + 6):

            obj_number = sheet_obj.cell(row=i, column=1)
            print(obj_number)
            obj_address = sheet_obj.cell(row=i, column=2)
            obj_comment = sheet_obj.cell(row=i, column=3)
            obj_dealer = sheet_obj.cell(row=i, column=4)
            new_obj , created = Object.objects.get_or_create(number=obj_number.value)
            if created:
                new_obj.address = obj_address.value
                new_obj.comment = obj_comment.value
                new_obj.save()

            new_obj.additional_equipments.all().delete()
            print(obj_number.value,obj_address.value,obj_comment.value,obj_dealer.value)
            for j in range(5, 45):
                print('id', sheet_obj.cell(row=3, column=j).value)
                print('amount', sheet_obj.cell(row=i, column=j).value)
                if sheet_obj.cell(row=i, column=j).value != ' ':
                    if sheet_obj.cell(row=i, column=j).value:
                        ObjectAdditionalEquipment.objects.create(
                                object=new_obj,
                                amount=sheet_obj.cell(row=i, column=j).value,
                                model_id=sheet_obj.cell(row=3, column=j).value)




        #     #Object.objects.create(number=obj_number.value,address=obj_address.value)

        return Response(status=200)

class ObjectUpdate(APIView):
    def post(self, request):
        obj = Object.objects.get(id=self.request.query_params.get('id'))

        setattr(request.data, '_mutable', True)

        try:
            request.data.pop('files')
            files_descriptions = request.data.pop('descriptions')
        except:
            files_descriptions = []
        data = json.loads(json.dumps(request.data))
        json_data = {}
        for dat in data:
            json_data[dat] = json.loads(data[dat])
        print(json_data['client'])
        serializer = ObjectSerializer(data=json_data,instance=obj)
        if serializer.is_valid():
            print('dd')
            serializer.save()
            obj.client_id = json_data['client']
            obj.save()

            for equipment in json_data['equipments']:
                if equipment['is_new']:
                    ObjectAdditionalEquipment.objects.create(
                        amount=equipment['amount'],
                        object=obj,
                        model_id=equipment['model']
                    )
            for contact in json_data['contacts']:
                if contact['is_new']:
                    ObjectContact.objects.create(
                        order_num=contact['order_num'],
                        object=obj,
                        name=contact['name'],
                        phone=contact['phone'],
                        email=contact['email'],
                        comment=contact['comment'],
                        social=contact['social'],
                    )
                else:
                    contact_obj = ObjectContact.objects.get(id=contact['id'])
                    contact_obj.order_num = contact['order_num']
                    contact_obj.name = contact['name']
                    contact_obj.phone = contact['phone']
                    contact_obj.email = contact['email']
                    contact_obj.comment = contact['comment']
                    contact_obj.social = contact['social']
                    contact_obj.save()


        else:
            print(serializer.errors)
        # order.object_id = json_data['object']
        # order.equipment_id = json_data['equipment']
        # type_id = json_data.get('type', None)
        # work_type_id = json_data.get('work_type', None)
        # if type_id:
        #     order.type_id = type_id
        # #if work_type_id:
        # order.work_type_id = work_type_id
        # # order.type_id = json_data['type']
        # # order.work_type_id = json_data['work_type']
        # order.is_critical = json_data['is_critical']
        # order.comment = json_data['comment']
        # order.date_dead_line = json_data['date_dead_line']
        # order.save()
        for index, file in enumerate(request.FILES.getlist('files')):
            ObjectFile.objects.create(file=file, object=obj, text=files_descriptions[index])

        return Response(status=200)