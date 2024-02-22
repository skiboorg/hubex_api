import json

from rest_framework.pagination import PageNumberPagination
from rest_framework.views import APIView
from rest_framework.response import Response
from .serializers import *
from .models import *
from rest_framework import generics, viewsets, parsers
import django_filters
from django_filters import IsoDateTimeFilter
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from chat.models import OrderChat
from user.models import User, UserWorkTime
from django.db.models import Count, Q
import datetime

from object.models import Object
from equipment.models import Equipment

from user.services import send_tg_mgs

class OrderPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 10000

class OrderFilter(django_filters.FilterSet):
    created_at_gte = IsoDateTimeFilter(field_name="date_created_at", lookup_expr='gte')
    created_at_lte = IsoDateTimeFilter(field_name="date_created_at", lookup_expr='lte')
    q = django_filters.CharFilter(method='my_custom_filter', label="Search")

    def my_custom_filter(self, queryset, name, value):
        return queryset.filter(
            Q(number__icontains=value) |
            Q(object__name__icontains=value) |
            Q(object__number__icontains=value) |
            Q(equipment__model__name__icontains=value) |
            Q(equipment__serial_number__icontains=value) |
            Q(object__address__icontains=value)
        )

    class Meta:
        model = Order
        fields = {
            'is_done': ('exact',),
            'is_critical': ('exact',),
            'status_id': ('in',),
        }
class OrderViewSet(viewsets.ModelViewSet):
    pagination_class = OrderPagination
    queryset = Order.objects.all()
    serializer_class = OrderSerializer
    lookup_field = 'number'
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = OrderFilter
    # filterset_fields = [
    #     'is_critical',
    #     'is_done',
    #     'created_at_gte',
    #     'created_at_lte',
    # ]

    # def perform_create(self, serializer):
    #     print('serializer.validated_data',serializer.validated_data)
    #     serializer.save()

    # def list(self, request, *args, **kwargs):
    #     qs = self.get_queryset()
    #
    #     for item in qs:
    #         if item.date_dead_line:
    #             if datetime.date.today() > item.date_dead_line:
    #                 item.is_time_left = True
    #                 item.save(update_fields=['is_time_left'])
    #     filtered_queryset = self.filter_queryset(qs)
    #     page = self.paginate_queryset(filtered_queryset)
    #     serializer = self.get_serializer(page, many=True)
    #     return Response(serializer.data)
    def get_serializer_class(self):
        full_mode = self.request.query_params.get('full', None)

        if full_mode:
            return OrderSerializer
        else:
            return OrderShortSerializer

    def create(self, request, *args, **kwargs):

        print(request.data)
        setattr(request.data, '_mutable', True)
        try:
            request.data.pop('files')
            files_descriptions = request.data.pop('descriptions')
        except:
            files_descriptions = []
        data = json.loads(json.dumps(request.data))
        json_data = {}
        for dat in data:
            json_data[dat] = json.loads(data[dat])
        serializer = self.get_serializer(data=json_data)
        if serializer.is_valid():
            obj = serializer.save()
            type_id = json_data.get('type', None)
            work_type_id = json_data.get('work_type', None)
            if type_id:
                obj.type_id = type_id
            if work_type_id:
                obj.work_type_id = work_type_id


            object_id = json_data['object']
            equipment_id = json_data['equipment']
            need_create_new_object = json_data.get('need_create_object',None)

            new_object = None
            if not need_create_new_object:
                obj.object_id = object_id
            else:
                new_object = Object.objects.create(
                     number= object_id.get('number'),
                    address=object_id.get('address'),
                    address_comment=object_id.get('address_comment'),
                 )
                obj.object = new_object


            if not need_create_new_object:
                if not equipment_id:
                    temp_equipment = Equipment.objects.get(is_temp_equipment=True)
                    obj.equipment = temp_equipment
                else:
                    obj.equipment_id = equipment_id
            else:
                new_equipment = Equipment.objects.create(
                    object=new_object,
                    serial_number=equipment_id.get('serial_number'),
                    model_id=equipment_id.get('model'),
                    name=equipment_id.get('name'),
                )
                obj.equipment = new_equipment
            if json_data['is_critical']:
                delta = 7
            else:
                delta = 14
            obj.date_dead_line = (datetime.datetime.now() + datetime.timedelta(days=delta)).date()
            obj.save()
            for index,file in enumerate(request.FILES.getlist('files')):
                OrderFile.objects.create(file=file,order=obj,text=files_descriptions[index])
            OrderChat.objects.create(order=obj)
        else:
            print(serializer.errors)

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    # def update(self, request, *args, **kwargs):
    #     print(self.request.data)

    def perform_update(self, serializer):
        stage_id = self.request.data.get('stage_id', None)
        is_done = self.request.data.get('is_done', None)

        if stage_id:
            new_stage = Stage.objects.get(id=stage_id)
            if new_stage.status:
                serializer.save(stage_id=stage_id, status=new_stage.status)
            else:
                serializer.save(stage_id=stage_id)

            StageLog.objects.create(user=self.request.user,order=serializer.instance,new_stage=new_stage)

        if is_done:
            done_status = Status.objects.get(is_done=True)
            serializer.save(is_done=True, status=done_status, date_done=(datetime.datetime.now()).date())

class SaveCheckListData(APIView):
    def post(self, request):
        print(request.data)
        obj, _ = CheckListData.objects.get_or_create(order_id=request.data['order_id'],check_list_id=request.data['check_list_id'])
        h_obj, _ = CheckListHistory.objects.get_or_create(order_id=request.data['order_id'],check_list_id=request.data['check_list_id'])
        CheckListDataHistory.objects.create(
            history_obj=h_obj,
            data=request.data['data'],
            created_at=datetime.datetime.now()
        )
        obj.data = request.data['data']
        obj.save()
        return Response(status=200)

class GetOrdersByWorker(generics.ListAPIView):
    serializer_class = OrdersForWorkerSerializer

    def get_queryset(self):
        user = self.request.user
        today_orders = []
        orders = Order.objects.filter(is_done=False, users__in=[user.id], stage__role__in=[user.role.id])
        if user.show_only_today_orders:
            times = UserWorkTime.objects.filter(date=str(datetime.datetime.today().date()))
            for time in times:
                order = orders.filter(id=time.order.id)
                if order.exists():
                    today_orders.append(order.first())
            return today_orders
        else:
            return orders

class GetOrdersByWorkerForCalendar(generics.ListAPIView):
    serializer_class = OrderSerializer

    def get_queryset(self):
        user_id = self.request.query_params.get('user_id')
        print(user_id)
        return Order.objects.filter(users__in=[user_id])
class GetOrdersHistoryByObject(generics.ListAPIView):
    serializer_class = OrderShortSerializer

    def get_queryset(self):
        return Order.objects.filter(equipment_id=self.kwargs.get('object_id'), is_done=True)


class GetOrdersByUser(generics.ListAPIView):
    serializer_class = OrderShortSerializer

    def get_queryset(self):
        print(self.kwargs.get('id'))
        return Order.objects.filter(users__in=[self.kwargs.get('id')])


class DeleteUserFromOrder(APIView):
    def post(self, request):
        order_uuid = request.data['order']
        user_id = request.data['user']
        order = Order.objects.get(uuid=order_uuid)
        user = User.objects.get(id=user_id)
        qs = UserWorkTime.objects.filter(order=order, user=user)
        order.users.remove(user)
        if qs.exists():
            qs.first().delete()
        return Response(status=200)



class AddUsersToOrder(APIView):

    def post(self, request):
        order_uuid = request.data['order']
        order_users = request.data['users']
        order = Order.objects.get(uuid=order_uuid)
        chat, _ = OrderChat.objects.get_or_create(order=order)
        print(request.data)

        staff_users = User.objects.filter(is_staff=True)
        for staff_user in staff_users:
            chat.users.add(staff_user)

        for order_user in order_users:
            user = User.objects.get(id=order_user['id'])
            order.users.add(user)
            chat.users.add(user)
            if order_user['events']:
                new_time = UserWorkTime.objects.create(
                    user=user,
                    order=order,
                    type_id=order_user['events']['type'],
                    date = order_user['events']['date'].replace('/','-'),
                    start_time=order_user['events']['start_time'],
                    end_time=order_user['events']['end_time'],
                )
                send_tg_mgs(user.telega_id, f'Вы назначены на заявку {order.number},дата {new_time.date}'
                                            f' c {new_time.start_time} до {new_time.end_time}')
            print(order_user)
        return Response(status=200)


class CheckOrders(APIView):
    def get(self,request):

        qs = Order.objects.filter(is_done=False)
        for item in qs:
            print(item)
            if item.date_dead_line:
                if datetime.date.today() > item.date_dead_line:
                    item.is_time_left = True
                    item.save(update_fields=['is_time_left'])
        return Response(status=200)

class AddUserToOrder(APIView):
    def post(self, request):
        print(request.data)
        order_uuid = request.data['order']
        order_user = request.data['user']
        order = Order.objects.get(uuid=order_uuid)
        chat, _ = OrderChat.objects.get_or_create(order=order)
        user = User.objects.get(uuid=order_user)
        order.users.add(user)
        chat.users.add(user)
        return Response(status=200)

class CheckListFilter(django_filters.FilterSet):

    q = django_filters.CharFilter(method='my_custom_filter', label="Search")

    def my_custom_filter(self, queryset, name, value):
        return queryset.filter(
            Q(check_list__name__icontains=value) |
            Q(order__number__icontains=value)
        )

    class Meta:
        model = CheckListData
        fields = ['created_at']


class GetCheckListsTemplates(generics.ListAPIView):
    serializer_class = CheckListSerializer
    queryset = CheckList.objects.all()



class GetCheckLists(generics.ListAPIView):
    pagination_class = OrderPagination
    serializer_class = CheckListDataShortSerializer
    queryset = CheckListData.objects.all()
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = CheckListFilter

class GetCheckList(generics.RetrieveAPIView):
    serializer_class = CheckListDataSerializer
    def get_object(self):
        return CheckListData.objects.get(id=self.request.query_params.get('id'))

class GetCheckListTemplate(generics.RetrieveAPIView):
    serializer_class = CheckListSerializer
    def get_object(self):
        return CheckList.objects.get(id=self.request.query_params.get('id'))

class OrderTypes(generics.ListAPIView):
    serializer_class = TypeSerializer
    queryset = Type.objects.all()

class OrderStatuses(generics.ListAPIView):
    serializer_class = StatusSerializer
    queryset = Status.objects.all()

class OrderWorkTypes(generics.ListAPIView):
    serializer_class = WorkTypeSerializer
    queryset = WorkType.objects.all()

class OrderDeleteFile(generics.DestroyAPIView):
    serializer_class = OrderFile
    queryset = OrderFile.objects.all()

class OrderAddFile(APIView):
    def post(self, request):
        order = Order.objects.get(number=self.request.query_params.get('number'))
        try:
            request.data.pop('files')
            files_descriptions = request.data.pop('descriptions')
        except:
            files_descriptions = []
        for index, file in enumerate(request.FILES.getlist('files')):
            OrderFile.objects.create(file=file, order=order, text=files_descriptions[index])
        return Response(status=200)
class OrderUpdate(APIView):
    def post(self, request):
        order = Order.objects.get(number=self.request.query_params.get('number'))
        print(request.data)
        setattr(request.data, '_mutable', True)

        try:
            request.data.pop('files')
            files_descriptions = request.data.pop('descriptions')
        except:
            files_descriptions = []
        data = json.loads(json.dumps(request.data))
        json_data = {}
        for dat in data:
            json_data[dat] = json.loads(data[dat])
        print(json_data)
        object_id = json_data.get('object', None)
        if object_id:
            order.object_id = object_id
        equipment_id = json_data.get('equipment',None)
        if not equipment_id:
            temp_equipment = Equipment.objects.get(is_temp_equipment=True)
            order.equipment = temp_equipment
        else:
            order.equipment_id = equipment_id

        type_id = json_data.get('type', None)
        work_type_id = json_data.get('work_type', None)
        if type_id:
            order.type_id = type_id
        #if work_type_id:
        order.work_type_id = work_type_id
        # order.type_id = json_data['type']
        # order.work_type_id = json_data['work_type']
        order.is_critical = json_data['is_critical']
        order.comment = json_data['comment']
        order.date_dead_line = json_data['date_dead_line']
        order.save()
        for index, file in enumerate(request.FILES.getlist('files')):
            OrderFile.objects.create(file=file, order=order, text=files_descriptions[index])

        return Response(status=200)



class OrderGetTableData(APIView):
    def get(self,request):
        order_id = self.request.query_params.get('order_id')
        check_list_id = self.request.query_params.get('check_list_id')
        table_id = self.request.query_params.get('table_id')
        try:
            data = CheckListTableData.objects.get(
                order_id=order_id,
                check_list_id=check_list_id,
                table_id=table_id
                                                  )
            serializer = CheckListTableDataSerializer(data)
            return Response({'data':serializer.data},status=200)
        except:
            return Response({'data':None},status=200)
class OrderUpdateChecklist(APIView):
    def post(self, request):
        data = request.data
        print(data['check_list'])
        print(data['tables'])
        check_list = CheckList.objects.get(id=data['check_list']['id'])
        inputs = check_list.inputs.all()
        inputs.delete()

        for input in data['check_list']['inputs']:
            labels = ''
            for label in input['labels']:
                labels += label['value'] + '/'
            CheckListInput.objects.create(
                check_list=check_list,
                input_id=input['input_type']['id'],
                label=input['label'],
                labels=labels[:-1]
            )
        for table in data['tables']:
            print(table)
            table_id = table.get('id', None)
            if not table_id:
                check_list_table = CheckListTable.objects.create(
                    name=table['name'],
                    check_list=check_list,
                    default_data=table['rows']
                )
            else:
                check_list_table = CheckListTable.objects.get(id=table_id)
                table_inputs = check_list_table.check_list_table_inputs.all()
                table_inputs.delete()
                check_list_table.name = table['name']
                check_list_table.default_data = table['rows']
                check_list_table.save()
            for input in table['inputs']:
                CheckListTableInput.objects.create(
                    table=check_list_table,
                    input_id=input['input']['id'],
                    label= input['label']
                )


        # tables = check_list.check_list_tables.all()
        # tables.delete()
        # for table in data['tables']:
        #     new_table = CheckListTable.objects.create(
        #         name=table['name'],
        #         check_list=check_list,
        #         default_data=table['rows']
        #     )
        #     for input in table['inputs']:
        #         CheckListTableInput.objects.create(
        #             table=new_table,
        #             input_id=input['input']['id'],
        #             label = input['label']
        #         )
        return Response(status=200)
class OrderCreateChecklist(APIView):
    def post(self, request):
        data = request.data
        print(data['tables'])
        new_check_list = CheckList.objects.create(
            name=data['check_list']['name']
        )

        for input in data['check_list']['inputs']:
            labels = ''
            for label in input['labels']:
                labels += label['value'] + '/'
            new_input = CheckListInput.objects.create(
                check_list=new_check_list,
                input_id=input['input_type']['id'],
                label=input['label'],
                labels=labels[:-1]
            )
        for table in data['tables']:
            new_table = CheckListTable.objects.create(
                name=table['name'],
                check_list=new_check_list,
                default_data=table['rows']
            )
            for input in table['inputs']:
                CheckListTableInput.objects.create(
                    table=new_table,
                    input_id=input['input_type']['id'],
                    label = input['label']
                )

        return Response(status=200)
class OrderSaveTable(APIView):
    def post(self, request):
        data = request.data
        table = CheckListTable.objects.get(id=data['table_id'])

        h_obj, _ = CheckListHistory.objects.get_or_create(order_id=data['order_id'],
                                                          check_list_id=data['check_list_id'])
        CheckListTableHistory.objects.create(
            history_obj=h_obj,
            data=data['data'],
            table=table,
            created_at=datetime.datetime.now()
        )

        if not table.default_data:
            table.default_data = data['data']
            table.save()
        else:
            table_data, _ = CheckListTableData.objects.get_or_create(
                order_id=data['order_id'],
                check_list_id=data['check_list_id'],
                table=table,
            )
            table_data.data = data['data']
            table_data.save()

        return Response(status=200)

class OrderGetChecklistInputs(generics.ListAPIView):
    serializer_class = InputFieldSerializer
    queryset = InputField.objects.all()

class OrderGetChecklistTableInputs(generics.ListAPIView):
    serializer_class = CheckListTableInputFieldSerializer
    queryset = CheckListTableInputField.objects.all()


class OrderFill(APIView):

    def get(self,request):
        from glob import glob

        from openpyxl import load_workbook
        wb = load_workbook(filename='orders.xlsx')
        sheet_obj = wb.active
        max_row = sheet_obj.max_row

        status_id = 4

        for i in range(1,  3):
            ord_number = sheet_obj.cell(row=i, column=1).value
            ord_date = sheet_obj.cell(row=i, column=2).value
            ord_comment = sheet_obj.cell(row=i, column=3).value
            ord_obj = sheet_obj.cell(row=i, column=4).value
            user_id = sheet_obj.cell(row=i, column=5).value
            user_date = sheet_obj.cell(row=i, column=6).value
            order, order_created = Order.objects.get_or_create(number=ord_number)
            if order_created:
                print('created new order')
                OrderChat.objects.create(order=order)
            print(order)
            print(ord_number,ord_date, ord_comment, ord_obj, user_id, user_date)

            obj = Object.objects.get(number=ord_obj)
            order.type_id = 1
            order.stage_id = 1
            order.comment = ord_comment
            order.object_id = obj.id
            order.date_created_at = ord_date
            order.status_id=status_id
            order.is_done = True
            order.users.add(user_id)
            order.equipment = obj.equipments.first()
            order.save()
            user_time, time_created = UserWorkTime.objects.get_or_create(order=order,
                                                                         date=user_date,user_id=user_id,
                                                                         start_time='09:00',end_time='21:00',type_id=1)

            print(obj.equipments.first())
            if order_created:
                dirs = glob("ord_files/*/", recursive = True)
                for dir in dirs:
                    if ord_number in dir:
                        print(dir)
                        files = glob(f"{dir}*", recursive = True)
                        print(files)
                        for file in files:
                            OrderFile.objects.create(order=order,file=file , text='Файл' )

        return Response(status=200)